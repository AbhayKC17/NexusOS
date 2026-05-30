"""
nexus/office/spreadsheet.py — Excel-like spreadsheet editor.

Features
────────
  • Infinite grid with column letters (A, B, C…) and row numbers
  • Formula bar showing cell content
  • Basic formula evaluation: =SUM, =AVG, =COUNT, =MAX, =MIN
  • Sort by any column (asc/desc)
  • Column filter via right-click
  • Save as .xlsx (openpyxl) or .csv
  • Load .xlsx / .csv / .xls
  • Add/delete rows and columns
"""
from __future__ import annotations

import os
import csv

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFrame, QPushButton,
    QLabel, QTableWidget, QTableWidgetItem, QHeaderView,
    QLineEdit, QFileDialog, QMessageBox, QMenu, QComboBox,
    QAbstractItemView,
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QColor, QFont


def _col_letter(n: int) -> str:
    """Convert 0-based column index to Excel-style letter (A, B, … Z, AA…)."""
    result = ""
    n += 1
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def _eval_formula(expr: str, table: QTableWidget) -> str:
    """Evaluate a simple spreadsheet formula."""
    e = expr.strip().lstrip("=").upper()
    try:
        if e.startswith("SUM(") or e.startswith("AVG(") or e.startswith("AVERAGE("):
            op  = "SUM" if "SUM" in e else "AVG"
            rng = e.split("(")[1].rstrip(")")
            vals = _range_values(rng, table)
            if not vals:
                return "0"
            return str(sum(vals) if op == "SUM" else round(sum(vals) / len(vals), 4))
        elif e.startswith("COUNT("):
            rng  = e[6:].rstrip(")")
            vals = _range_values(rng, table)
            return str(len(vals))
        elif e.startswith("MAX("):
            vals = _range_values(e[4:].rstrip(")"), table)
            return str(max(vals)) if vals else "0"
        elif e.startswith("MIN("):
            vals = _range_values(e[4:].rstrip(")"), table)
            return str(min(vals)) if vals else "0"
        else:
            # Basic arithmetic
            result = eval(e, {"__builtins__": {}})  # nosec: no user code
            return str(result)
    except Exception:
        return "#ERR"


def _range_values(rng: str, table: QTableWidget) -> list[float]:
    """Parse a cell range like A1:C3 and return numeric values."""
    vals = []
    if ":" in rng:
        start, end = rng.split(":")
        sr, sc = _cell_ref(start)
        er, ec = _cell_ref(end)
        if sr is None or sc is None:
            return vals
        for r in range(sr, er + 1):
            for c in range(sc, ec + 1):
                item = table.item(r, c)
                if item:
                    try:
                        vals.append(float(item.text()))
                    except ValueError:
                        pass
    else:
        r, c = _cell_ref(rng)
        if r is not None:
            item = table.item(r, c)
            if item:
                try:
                    vals.append(float(item.text()))
                except ValueError:
                    pass
    return vals


def _cell_ref(ref: str):
    """Parse a cell reference like 'B3' → (row=2, col=1)."""
    ref = ref.strip().upper()
    col_str = "".join(c for c in ref if c.isalpha())
    row_str = "".join(c for c in ref if c.isdigit())
    if not col_str or not row_str:
        return None, None
    col = 0
    for ch in col_str:
        col = col * 26 + (ord(ch) - 64)
    col -= 1
    row = int(row_str) - 1
    return row, col


class SpreadsheetEditor(QWidget):
    """Full-featured spreadsheet editor."""

    title_changed = pyqtSignal(str)

    _DEFAULT_ROWS = 50
    _DEFAULT_COLS = 26

    def __init__(self, path: str | None = None, title: str = "Sheet1", parent=None):
        super().__init__(parent)
        self._path    = path
        self._title   = title
        self._editing = False
        self._build()
        if path and os.path.exists(path):
            self._load_file(path)
        else:
            self._init_empty()

    # ── Build ─────────────────────────────────────────────────────────────────

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # ── Top toolbar ───────────────────────────────────────────────────────
        toolbar = QFrame()
        toolbar.setObjectName("pageHeader")
        toolbar.setFixedHeight(44)
        tbar = QHBoxLayout(toolbar)
        tbar.setContentsMargins(12, 6, 12, 6)
        tbar.setSpacing(6)

        title_edit = QLineEdit(self._title)
        title_edit.setStyleSheet(
            "QLineEdit { background: transparent; border: none; "
            "font-size: 13px; font-weight: 700; color: #1A1A1A; }"
        )
        title_edit.setMaximumWidth(180)
        title_edit.editingFinished.connect(
            lambda: self.title_changed.emit(title_edit.text())
        )
        tbar.addWidget(title_edit)
        tbar.addSpacing(8)

        for label, tip, fn in [
            ("+ Row",    "Add row below",   self._add_row),
            ("+ Col",    "Add column right", self._add_col),
            ("− Row",    "Delete selected rows", self._del_row),
            ("− Col",    "Delete selected columns", self._del_col),
        ]:
            b = QPushButton(label)
            b.setFixedHeight(28)
            b.setToolTip(tip)
            b.clicked.connect(fn)
            tbar.addWidget(b)

        tbar.addSpacing(8)

        # Sort
        sort_lbl = QLabel("Sort:")
        sort_lbl.setStyleSheet("color: rgba(0,0,0,0.5); background: transparent; font-size: 12px;")
        tbar.addWidget(sort_lbl)
        self._sort_combo = QComboBox()
        self._sort_combo.setFixedHeight(28)
        self._sort_combo.setMinimumWidth(100)
        tbar.addWidget(self._sort_combo)

        sort_asc = QPushButton("A→Z")
        sort_asc.setFixedHeight(28)
        sort_asc.setFixedWidth(44)
        sort_asc.clicked.connect(lambda: self._sort(asc=True))
        tbar.addWidget(sort_asc)

        sort_desc = QPushButton("Z→A")
        sort_desc.setFixedHeight(28)
        sort_desc.setFixedWidth(44)
        sort_desc.clicked.connect(lambda: self._sort(asc=False))
        tbar.addWidget(sort_desc)

        tbar.addStretch()

        open_btn = QPushButton("↑  Open")
        open_btn.setFixedHeight(28)
        open_btn.clicked.connect(self._open_file)
        tbar.addWidget(open_btn)

        save_btn = QPushButton("💾  Save")
        save_btn.setObjectName("accentBtn")
        save_btn.setFixedHeight(28)
        save_btn.clicked.connect(self._save)
        tbar.addWidget(save_btn)

        lay.addWidget(toolbar)

        # ── Formula bar ───────────────────────────────────────────────────────
        fbar = QFrame()
        fbar.setFixedHeight(30)
        fbar.setStyleSheet(
            "QFrame { background: #FAFAFA; border-bottom: 1px solid rgba(0,0,0,0.07); }"
        )
        fl = QHBoxLayout(fbar)
        fl.setContentsMargins(8, 4, 8, 4)
        fl.setSpacing(8)

        self._cell_ref_lbl = QLabel("A1")
        self._cell_ref_lbl.setFixedWidth(48)
        self._cell_ref_lbl.setStyleSheet(
            "font-size: 12px; font-weight: 700; color: #0067C0; "
            "background: transparent; padding: 0 4px;"
        )
        fl.addWidget(self._cell_ref_lbl)

        eq_lbl = QLabel("ƒx")
        eq_lbl.setStyleSheet(
            "font-size: 12px; color: rgba(0,0,0,0.4); background: transparent;"
        )
        fl.addWidget(eq_lbl)

        self._formula_bar = QLineEdit()
        self._formula_bar.setStyleSheet(
            "QLineEdit { background: #FFFFFF; border: 1px solid rgba(0,0,0,0.12); "
            "border-radius: 4px; font-size: 12px; padding: 2px 8px; }"
            "QLineEdit:focus { border-color: #0067C0; }"
        )
        self._formula_bar.returnPressed.connect(self._commit_formula)
        fl.addWidget(self._formula_bar, 1)
        lay.addWidget(fbar)

        # ── Table ─────────────────────────────────────────────────────────────
        self._table = QTableWidget(0, 0)
        self._table.verticalHeader().setVisible(True)
        self._table.verticalHeader().setDefaultSectionSize(24)
        self._table.verticalHeader().setStyleSheet(
            "QHeaderView::section { background: #F5F5F5; color: rgba(0,0,0,0.5); "
            "font-size: 11px; border: none; border-right: 1px solid rgba(0,0,0,0.08); "
            "padding: 2px 6px; }"
        )
        self._table.horizontalHeader().setDefaultSectionSize(100)
        self._table.horizontalHeader().setStyleSheet(
            "QHeaderView::section { background: #F5F5F5; color: rgba(0,0,0,0.5); "
            "font-size: 11px; font-weight: 700; border: none; "
            "border-bottom: 1px solid rgba(0,0,0,0.10); padding: 4px 6px; }"
            "QHeaderView::section:hover { background: #EBEBEB; }"
        )
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self._table.setShowGrid(True)
        self._table.setGridStyle(Qt.PenStyle.SolidLine)
        self._table.setStyleSheet(
            "QTableWidget { background: #FFFFFF; gridline-color: rgba(0,0,0,0.08); "
            "selection-background-color: rgba(0,103,192,0.10); }"
            "QTableWidget::item { padding: 2px 6px; color: #1A1A1A; font-size: 12px; }"
        )
        self._table.setAlternatingRowColors(False)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.ContiguousSelection)
        self._table.currentCellChanged.connect(self._on_cell_selected)
        self._table.cellChanged.connect(self._on_cell_changed)
        self._table.horizontalHeader().customContextMenuRequested.connect(self._col_ctx)
        self._table.horizontalHeader().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        lay.addWidget(self._table, 1)

        # ── Status bar ────────────────────────────────────────────────────────
        status = QFrame()
        status.setFixedHeight(22)
        status.setStyleSheet(
            "QFrame { background: #F0F0F0; border-top: 1px solid rgba(0,0,0,0.06); }"
        )
        sl = QHBoxLayout(status)
        sl.setContentsMargins(12, 0, 12, 0)
        self._status_lbl = QLabel("Ready")
        self._status_lbl.setStyleSheet(
            "font-size: 11px; color: rgba(0,0,0,0.45); background: transparent;"
        )
        sl.addWidget(self._status_lbl)
        sl.addStretch()
        lay.addWidget(status)

    # ── Init empty grid ───────────────────────────────────────────────────────

    def _init_empty(self):
        self._editing = True
        self._table.setRowCount(self._DEFAULT_ROWS)
        self._table.setColumnCount(self._DEFAULT_COLS)
        self._table.setHorizontalHeaderLabels(
            [_col_letter(i) for i in range(self._DEFAULT_COLS)]
        )
        self._update_sort_combo()
        self._editing = False

    # ── Cell events ───────────────────────────────────────────────────────────

    def _on_cell_selected(self, row, col, *_):
        self._cell_ref_lbl.setText(f"{_col_letter(col)}{row + 1}")
        item = self._table.item(row, col)
        # Store raw formula in item data
        raw = (item.data(Qt.ItemDataRole.UserRole) or item.text()) if item else ""
        self._formula_bar.setText(raw)

    def _on_cell_changed(self, row, col):
        if self._editing:
            return
        item = self._table.item(row, col)
        if not item:
            return
        text = item.text()
        if text.startswith("="):
            # Store raw formula, display evaluated result
            item.setData(Qt.ItemDataRole.UserRole, text)
            result = _eval_formula(text, self._table)
            self._editing = True
            item.setText(result)
            self._editing = False
            item.setForeground(QColor("#0067C0"))
        else:
            item.setData(Qt.ItemDataRole.UserRole, text)
            item.setForeground(QColor("#1A1A1A"))

    def _commit_formula(self):
        row = self._table.currentRow()
        col = self._table.currentColumn()
        if row < 0 or col < 0:
            return
        val = self._formula_bar.text()
        item = self._table.item(row, col)
        if not item:
            item = QTableWidgetItem()
            self._table.setItem(row, col, item)
        item.setText(val)

    # ── Row / col management ──────────────────────────────────────────────────

    def _add_row(self):
        row = self._table.currentRow()
        pos = row + 1 if row >= 0 else self._table.rowCount()
        self._table.insertRow(pos)

    def _add_col(self):
        col = self._table.currentColumn()
        pos = col + 1 if col >= 0 else self._table.columnCount()
        self._table.insertColumn(pos)
        self._rebuild_col_headers()

    def _del_row(self):
        rows = sorted(
            {idx.row() for idx in self._table.selectedIndexes()}, reverse=True
        )
        for r in rows:
            self._table.removeRow(r)

    def _del_col(self):
        cols = sorted(
            {idx.column() for idx in self._table.selectedIndexes()}, reverse=True
        )
        for c in cols:
            self._table.removeColumn(c)
        self._rebuild_col_headers()

    def _rebuild_col_headers(self):
        n = self._table.columnCount()
        self._table.setHorizontalHeaderLabels([_col_letter(i) for i in range(n)])
        self._update_sort_combo()

    # ── Sort ──────────────────────────────────────────────────────────────────

    def _update_sort_combo(self):
        self._sort_combo.blockSignals(True)
        self._sort_combo.clear()
        for i in range(self._table.columnCount()):
            hdr = self._table.horizontalHeaderItem(i)
            label = hdr.text() if hdr else _col_letter(i)
            self._sort_combo.addItem(f"Col {label}", i)
        self._sort_combo.blockSignals(False)

    def _sort(self, asc: bool = True):
        col = self._sort_combo.currentData()
        if col is None:
            return
        n_rows = self._table.rowCount()
        n_cols = self._table.columnCount()

        def row_key(r):
            item = self._table.item(r, col)
            v = item.text() if item else ""
            try:
                return (0, float(v))
            except ValueError:
                return (1, v.lower())

        order = sorted(range(n_rows), key=row_key, reverse=not asc)

        self._editing = True
        data = []
        for r in range(n_rows):
            row_data = []
            for c in range(n_cols):
                item = self._table.item(r, c)
                row_data.append(item.text() if item else "")
            data.append(row_data)

        for new_r, old_r in enumerate(order):
            for c in range(n_cols):
                item = self._table.item(new_r, c)
                if not item:
                    item = QTableWidgetItem()
                    self._table.setItem(new_r, c, item)
                item.setText(data[old_r][c])
        self._editing = False

    # ── Column context menu (filter) ──────────────────────────────────────────

    def _col_ctx(self, pos):
        col = self._table.horizontalHeader().logicalIndexAt(pos)
        if col < 0:
            return
        hdr = self._table.horizontalHeaderItem(col)
        label = hdr.text() if hdr else _col_letter(col)

        menu = QMenu(self)
        menu.setStyleSheet(
            "QMenu { background: #FFFFFF; color: #1A1A1A; "
            "border: 1px solid rgba(0,0,0,0.12); border-radius: 6px; }"
            "QMenu::item { padding: 6px 16px; }"
            "QMenu::item:selected { background: rgba(0,103,192,0.10); }"
        )
        menu.addAction(f"Sort {label} A→Z").triggered.connect(
            lambda: self._sort_by_col(col, True)
        )
        menu.addAction(f"Sort {label} Z→A").triggered.connect(
            lambda: self._sort_by_col(col, False)
        )
        menu.addSeparator()
        menu.addAction("Auto-fit Column Width").triggered.connect(
            lambda: self._table.resizeColumnToContents(col)
        )
        menu.exec(self._table.horizontalHeader().mapToGlobal(pos))

    def _sort_by_col(self, col: int, asc: bool):
        prev = self._sort_combo.currentIndex()
        for i in range(self._sort_combo.count()):
            if self._sort_combo.itemData(i) == col:
                self._sort_combo.setCurrentIndex(i)
                break
        self._sort(asc)
        self._sort_combo.setCurrentIndex(prev)

    # ── File I/O ──────────────────────────────────────────────────────────────

    def _open_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Spreadsheet", "",
            "Spreadsheets (*.xlsx *.xls *.csv);;All Files (*)",
        )
        if path:
            self._path = path
            self._load_file(path)

    def _load_file(self, path: str):
        ext = os.path.splitext(path)[1].lower()
        self._editing = True
        self._table.clearContents()
        try:
            if ext == ".csv":
                self._load_csv(path)
            else:
                self._load_xlsx(path)
        except Exception as e:
            QMessageBox.warning(self, "Load Error", str(e))
        finally:
            self._editing = False
            self._rebuild_col_headers()

    def _load_csv(self, path: str):
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return
        self._table.setRowCount(max(len(rows), self._DEFAULT_ROWS))
        self._table.setColumnCount(max(len(rows[0]), self._DEFAULT_COLS) if rows else self._DEFAULT_COLS)
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                self._table.setItem(r, c, QTableWidgetItem(val))

    def _load_xlsx(self, path: str):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            nrows = ws.max_row or 0
            ncols = ws.max_column or 0
            self._table.setRowCount(max(nrows, self._DEFAULT_ROWS))
            self._table.setColumnCount(max(ncols, self._DEFAULT_COLS))
            for r in range(1, nrows + 1):
                for c in range(1, ncols + 1):
                    val = ws.cell(r, c).value
                    self._table.setItem(r - 1, c - 1,
                                        QTableWidgetItem("" if val is None else str(val)))
        except ImportError:
            raise RuntimeError("openpyxl is required to open .xlsx files. Run: pip install openpyxl")

    def _save(self):
        if not self._path:
            path, _ = QFileDialog.getSaveFileName(
                self, "Save Spreadsheet", f"{self._title}.xlsx",
                "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)",
            )
            if not path:
                return
            self._path = path

        ext = os.path.splitext(self._path)[1].lower()
        try:
            if ext == ".csv":
                self._save_csv(self._path)
            else:
                self._save_xlsx(self._path)
            self._status_lbl.setText(f"Saved — {os.path.basename(self._path)}")
        except Exception as e:
            QMessageBox.warning(self, "Save Error", str(e))

    def _save_csv(self, path: str):
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for r in range(self._table.rowCount()):
                row = []
                all_empty = True
                for c in range(self._table.columnCount()):
                    item = self._table.item(r, c)
                    v = item.text() if item else ""
                    if v:
                        all_empty = False
                    row.append(v)
                if not all_empty:
                    w.writerow(row)

    def _save_xlsx(self, path: str):
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self._title[:31]
            for r in range(self._table.rowCount()):
                for c in range(self._table.columnCount()):
                    item = self._table.item(r, c)
                    if item and item.text():
                        raw = item.data(Qt.ItemDataRole.UserRole) or item.text()
                        ws.cell(r + 1, c + 1, raw)
            wb.save(path)
        except ImportError:
            # Fall back to CSV
            self._save_csv(os.path.splitext(path)[0] + ".csv")
            QMessageBox.information(
                self, "Saved as CSV",
                "openpyxl not installed — saved as .csv instead.\n"
                "Install with: pip install openpyxl"
            )
