"""
nexus/viewers.py — Right-panel viewers for NexusOS.

Classes
───────
  NodeInspector  — editable node details + edge list
  ExcelViewer    — openpyxl-backed spreadsheet editor
  TextViewer     — code / plain-text editor with VS Code launch
  ImageViewer    — zoomable image display
  NoteEditor     — rich text note that persists to the graph DB
"""
from __future__ import annotations

import os
import subprocess

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QTextEdit, QScrollArea,
    QLineEdit, QFrame, QSizePolicy, QHeaderView, QMessageBox,
    QStackedWidget,
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QFont, QPixmap

from nexus.graph_db import update_node, edges_for_node, get_node


# ── App Viewer (embeds full apps like JobTracker inside NexusOS) ──────────────

class AppViewer(QWidget):
    """
    Embeds a full app as a first-class tab inside NexusOS.

    • Shows the app's UI directly — no separate window needed.
    • "⤢ Pop out" button detaches it into its own window for more space.
    """

    def __init__(self, node_data: dict, parent=None):
        super().__init__(parent)
        self._node = node_data
        self._detached: "QMainWindow | None" = None
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # ── App title bar ─────────────────────────────────────────────────────
        bar = QFrame()
        bar.setObjectName("pageHeader")
        bar.setFixedHeight(44)
        bar_lay = QHBoxLayout(bar)
        bar_lay.setContentsMargins(14, 6, 12, 6)
        bar_lay.setSpacing(10)

        icon_lbl = QLabel("⊡")
        icon_lbl.setStyleSheet(
            "color: #7C3AED; font-size: 16px; font-weight: 700; background: transparent;"
        )
        bar_lay.addWidget(icon_lbl)

        title_lbl = QLabel(self._node.get("label", "App"))
        title_lbl.setStyleSheet(
            "font-size: 14px; font-weight: 700; background: transparent; color: #1A1A1A;"
        )
        bar_lay.addWidget(title_lbl)

        summary = (self._node.get("summary") or "")[:80]
        if summary:
            sub = QLabel(f"—  {summary}")
            sub.setStyleSheet(
                "font-size: 11px; color: rgba(0,0,0,0.42); background: transparent;"
            )
            bar_lay.addWidget(sub)

        bar_lay.addStretch()

        popout_btn = QPushButton("⤢  Pop out")
        popout_btn.setFixedHeight(28)
        popout_btn.setToolTip("Open this app in its own window")
        popout_btn.clicked.connect(self._pop_out)
        bar_lay.addWidget(popout_btn)

        lay.addWidget(bar)

        # ── Embedded app content ──────────────────────────────────────────────
        self._content_stack = QStackedWidget()
        self._content_stack.addWidget(self._build_app_widget())
        lay.addWidget(self._content_stack, 1)

    def _build_app_widget(self) -> QWidget:
        label  = self._node.get("label", "")
        meta   = self._node.get("meta") or {}

        is_jt = "JobTracker" in label or meta.get("app_module") == "shell.window"

        if is_jt:
            try:
                from shell.window import MainWindow
                w = MainWindow()
                # Embed as a widget (removes the OS window frame)
                w.setWindowFlag(Qt.WindowType.Widget, True)
                w.statusBar().hide()   # NexusOS has its own status bar
                return w
            except Exception as e:
                return self._placeholder(f"Could not load JobTracker:\n{e}")

        return self._placeholder(
            f"'{label}' does not have an embedded viewer yet.\n"
            "Use 'Pop out' to open it, or double-click its folder node."
        )

    @staticmethod
    def _placeholder(msg: str) -> QLabel:
        lbl = QLabel(msg)
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl.setWordWrap(True)
        lbl.setStyleSheet(
            "color: rgba(0,0,0,0.38); font-size: 13px; background: #F4F4FA; padding: 40px;"
        )
        return lbl

    def _pop_out(self):
        label = self._node.get("label", "App")
        meta  = self._node.get("meta") or {}
        if "JobTracker" in label or meta.get("app_module") == "shell.window":
            try:
                from shell.window import MainWindow
                w = MainWindow()
                w.show()
                self._detached = w
            except Exception as e:
                QMessageBox.warning(self, "Pop out failed", str(e))


# ── Node Inspector ────────────────────────────────────────────────────────────

class NodeInspector(QWidget):
    """Detailed, editable view of a single graph node."""

    open_file_requested = pyqtSignal(dict)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._node: dict | None = None
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(10)

        # ── Header row: type badge + Open button ──────────────────────────────
        header = QHBoxLayout()
        self._type_badge = QLabel("—")
        self._type_badge.setFixedHeight(22)
        self._type_badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._type_badge.setStyleSheet(
            "border-radius: 4px; padding: 2px 8px; "
            "font-size: 11px; font-weight: 700;"
        )
        header.addWidget(self._type_badge)
        header.addStretch()

        self._open_btn = QPushButton("Open →")
        self._open_btn.setObjectName("accentBtn")
        self._open_btn.setFixedHeight(28)
        self._open_btn.hide()
        self._open_btn.clicked.connect(self._on_open)
        header.addWidget(self._open_btn)
        root.addLayout(header)

        # ── Label ─────────────────────────────────────────────────────────────
        self._label_edit = QLineEdit()
        self._label_edit.setPlaceholderText("Label")
        self._label_edit.setStyleSheet("font-size: 15px; font-weight: 700;")
        self._label_edit.editingFinished.connect(self._save)
        root.addWidget(self._label_edit)

        # ── Summary ───────────────────────────────────────────────────────────
        summary_hdr = QLabel("SUMMARY")
        summary_hdr.setObjectName("sectionHeader")
        root.addWidget(summary_hdr)

        self._summary_edit = QTextEdit()
        self._summary_edit.setPlaceholderText("Describe this node…")
        self._summary_edit.setMaximumHeight(88)
        # Save on focus-out via monkey-patch
        orig_focus_out = self._summary_edit.focusOutEvent
        def _summary_focus_out(e):
            self._save()
            orig_focus_out(e)
        self._summary_edit.focusOutEvent = _summary_focus_out
        root.addWidget(self._summary_edit)

        # ── Path ─────────────────────────────────────────────────────────────
        self._path_lbl = QLabel()
        self._path_lbl.setStyleSheet(
            "color: rgba(239,239,239,0.38); font-size: 11px; background: transparent;"
        )
        self._path_lbl.setWordWrap(True)
        root.addWidget(self._path_lbl)

        # ── Connections ───────────────────────────────────────────────────────
        conn_hdr = QLabel("CONNECTIONS")
        conn_hdr.setObjectName("sectionHeader")
        root.addWidget(conn_hdr)

        self._conn_container = QWidget()
        self._conn_lay = QVBoxLayout(self._conn_container)
        self._conn_lay.setContentsMargins(0, 0, 0, 0)
        self._conn_lay.setSpacing(3)
        root.addWidget(self._conn_container)

        root.addStretch()

        # ── Empty state ───────────────────────────────────────────────────────
        self._empty = QLabel("← Select a node\nto inspect")
        self._empty.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._empty.setStyleSheet(
            "color: rgba(239,239,239,0.24); font-size: 13px; "
            "background: transparent; line-height: 1.6;"
        )
        root.addWidget(self._empty)

    # ── Public API ────────────────────────────────────────────────────────────

    def load_node(self, data: dict):
        self._node = data
        self._empty.hide()

        # Type badge
        try:
            from nexus.canvas import _TYPE
            color = _TYPE.get(data.get("type", "DEFAULT"), ("#6B7280", "○"))[0]
        except Exception:
            color = "#6B7280"

        t = data.get("type", "DEFAULT")
        self._type_badge.setText(t)
        self._type_badge.setStyleSheet(
            f"border-radius: 4px; padding: 2px 8px; font-size: 11px; font-weight: 700; "
            f"background: {color}30; color: {color}; border: 1px solid {color}55;"
        )

        self._label_edit.setText(data.get("label", ""))
        self._summary_edit.setPlainText(data.get("summary", ""))

        path = data.get("path") or ""
        self._path_lbl.setText(path)
        self._path_lbl.setVisible(bool(path))

        opens = t in ("FILE_EXCEL", "FILE_PDF", "FILE_TEXT", "FILE_CODE", "FILE_IMAGE", "NOTE")
        self._open_btn.setVisible(opens)

        # Rebuild connections list
        for i in reversed(range(self._conn_lay.count())):
            w = self._conn_lay.itemAt(i).widget()
            if w:
                w.setParent(None)

        for edge in edges_for_node(data["id"])[:8]:
            other_id = edge["tgt_id"] if edge["src_id"] == data["id"] else edge["src_id"]
            other = get_node(other_id)
            if not other:
                continue
            direction = "→" if edge["src_id"] == data["id"] else "←"
            txt = f"{direction} {other['label']}"
            if edge.get("label"):
                txt += f"  ·  {edge['label']}"
            pill = QLabel(txt)
            pill.setStyleSheet(
                "background: rgba(255,255,255,0.05); border-radius: 5px; "
                "padding: 3px 8px; color: rgba(239,239,239,0.60); font-size: 11px;"
            )
            self._conn_lay.addWidget(pill)

    def clear(self):
        self._node = None
        self._label_edit.clear()
        self._summary_edit.clear()
        self._path_lbl.clear()
        self._open_btn.hide()
        self._type_badge.setText("—")
        for i in reversed(range(self._conn_lay.count())):
            w = self._conn_lay.itemAt(i).widget()
            if w:
                w.setParent(None)
        self._empty.show()

    # ── Private ───────────────────────────────────────────────────────────────

    def _save(self):
        if not self._node:
            return
        update_node(
            self._node["id"],
            label=self._label_edit.text().strip() or self._node["label"],
            summary=self._summary_edit.toPlainText(),
        )

    def _on_open(self):
        if self._node:
            self.open_file_requested.emit(self._node)


# ── Excel Viewer / Editor ─────────────────────────────────────────────────────

class ExcelViewer(QWidget):
    """Editable spreadsheet view backed by openpyxl."""

    def __init__(self, path: str, parent=None):
        super().__init__(parent)
        self._path = path
        self._wb = None
        self._ws = None
        self._build()
        self._load()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # Toolbar
        bar = QFrame()
        bar.setObjectName("card")
        bar.setFixedHeight(44)
        bar_lay = QHBoxLayout(bar)
        bar_lay.setContentsMargins(12, 6, 12, 6)
        bar_lay.setSpacing(8)

        name_lbl = QLabel(os.path.basename(self._path))
        name_lbl.setStyleSheet(
            "color: rgba(239,239,239,0.65); font-size: 12px; background: transparent;"
        )
        bar_lay.addWidget(name_lbl)
        bar_lay.addStretch()

        save_btn = QPushButton("Save")
        save_btn.setObjectName("accentBtn")
        save_btn.setFixedHeight(28)
        save_btn.clicked.connect(self._save)
        bar_lay.addWidget(save_btn)
        lay.addWidget(bar)

        self._table = QTableWidget()
        self._table.setAlternatingRowColors(True)
        self._table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive
        )
        lay.addWidget(self._table)

    def _load(self):
        try:
            import openpyxl
            self._wb = openpyxl.load_workbook(self._path, data_only=True)
            self._ws = self._wb.active
            nrows = self._ws.max_row or 0
            ncols = self._ws.max_column or 0
            self._table.setRowCount(nrows)
            self._table.setColumnCount(ncols)
            for r in range(1, nrows + 1):
                for c in range(1, ncols + 1):
                    val = self._ws.cell(r, c).value
                    self._table.setItem(r - 1, c - 1,
                                        QTableWidgetItem("" if val is None else str(val)))
        except ImportError:
            self._table.setRowCount(1)
            self._table.setColumnCount(1)
            self._table.setItem(0, 0, QTableWidgetItem("openpyxl not installed"))
        except Exception as e:
            self._table.setRowCount(1)
            self._table.setColumnCount(1)
            self._table.setItem(0, 0, QTableWidgetItem(f"Error: {e}"))

    def _save(self):
        if not self._wb or not self._ws:
            return
        try:
            for r in range(self._table.rowCount()):
                for c in range(self._table.columnCount()):
                    item = self._table.item(r, c)
                    self._ws.cell(r + 1, c + 1, item.text() if item else "")
            self._wb.save(self._path)
            QMessageBox.information(self, "Saved", f"Saved — {os.path.basename(self._path)}")
        except Exception as e:
            QMessageBox.warning(self, "Save Error", str(e))


# ── Text / Code Viewer ────────────────────────────────────────────────────────

class TextViewer(QWidget):
    """Editable text / code viewer with VS Code launch."""

    def __init__(self, path: str, parent=None):
        super().__init__(parent)
        self._path = path
        self._build()
        self._load()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        bar = QFrame()
        bar.setObjectName("card")
        bar.setFixedHeight(44)
        bar_lay = QHBoxLayout(bar)
        bar_lay.setContentsMargins(12, 6, 12, 6)
        bar_lay.setSpacing(8)

        name_lbl = QLabel(os.path.basename(self._path))
        name_lbl.setStyleSheet(
            "color: rgba(239,239,239,0.65); font-size: 12px; background: transparent;"
        )
        bar_lay.addWidget(name_lbl)
        bar_lay.addStretch()

        vscode_btn = QPushButton("Open in VS Code")
        vscode_btn.setFixedHeight(28)
        vscode_btn.clicked.connect(self._open_vscode)
        bar_lay.addWidget(vscode_btn)

        save_btn = QPushButton("Save")
        save_btn.setObjectName("accentBtn")
        save_btn.setFixedHeight(28)
        save_btn.clicked.connect(self._save)
        bar_lay.addWidget(save_btn)
        lay.addWidget(bar)

        self._editor = QTextEdit()
        self._editor.setFont(QFont("Menlo, SF Mono, Consolas, monospace", 12))
        lay.addWidget(self._editor)

    def _load(self):
        try:
            with open(self._path, "r", encoding="utf-8", errors="replace") as fh:
                self._editor.setPlainText(fh.read())
        except Exception as e:
            self._editor.setPlainText(f"# Error loading file\n# {e}")

    def _save(self):
        try:
            with open(self._path, "w", encoding="utf-8") as fh:
                fh.write(self._editor.toPlainText())
        except Exception as e:
            QMessageBox.warning(self, "Save Error", str(e))

    def _open_vscode(self):
        try:
            subprocess.Popen(["code", self._path])
        except FileNotFoundError:
            QMessageBox.information(
                self, "VS Code",
                "Install VS Code and add 'code' to your PATH, then try again."
            )


# ── Image Viewer ──────────────────────────────────────────────────────────────

class ImageViewer(QWidget):
    """Zoomable image display."""

    def __init__(self, path: str, parent=None):
        super().__init__(parent)
        self._path = path
        self._zoom = 1.0
        self._pixmap = QPixmap()
        self._build()
        self._load()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        bar = QFrame()
        bar.setObjectName("card")
        bar.setFixedHeight(44)
        bar_lay = QHBoxLayout(bar)
        bar_lay.setContentsMargins(12, 6, 12, 6)
        bar_lay.setSpacing(6)

        name_lbl = QLabel(os.path.basename(self._path))
        name_lbl.setStyleSheet(
            "color: rgba(239,239,239,0.65); font-size: 12px; background: transparent;"
        )
        bar_lay.addWidget(name_lbl)
        bar_lay.addStretch()

        for symbol, delta in (("−", -0.2), ("+", 0.2)):
            btn = QPushButton(symbol)
            btn.setFixedSize(30, 28)
            btn.clicked.connect(lambda _, d=delta: self._zoom_by(d))
            bar_lay.addWidget(btn)

        fit_btn = QPushButton("Fit")
        fit_btn.setFixedHeight(28)
        fit_btn.clicked.connect(self._reset_zoom)
        bar_lay.addWidget(fit_btn)
        lay.addWidget(bar)

        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(False)
        self._scroll.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._scroll.setStyleSheet("background: #07070C; border: none;")

        self._img_lbl = QLabel()
        self._img_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._scroll.setWidget(self._img_lbl)
        lay.addWidget(self._scroll)

    def _load(self):
        self._pixmap = QPixmap(self._path)
        self._apply_zoom()

    def _apply_zoom(self):
        if self._pixmap.isNull():
            self._img_lbl.setText("Cannot load image")
            return
        scaled = self._pixmap.scaled(
            int(self._pixmap.width() * self._zoom),
            int(self._pixmap.height() * self._zoom),
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation,
        )
        self._img_lbl.setPixmap(scaled)
        self._img_lbl.resize(scaled.size())

    def _zoom_by(self, delta: float):
        self._zoom = max(0.1, min(4.0, self._zoom + delta))
        self._apply_zoom()

    def _reset_zoom(self):
        self._zoom = 1.0
        self._apply_zoom()


# ── Note Editor ───────────────────────────────────────────────────────────────

class NoteEditor(QWidget):
    """Rich text note editor that persists content to the graph DB."""

    def __init__(self, node_data: dict, parent=None):
        super().__init__(parent)
        self._node = node_data
        self._build()
        self._load()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        bar = QFrame()
        bar.setObjectName("card")
        bar.setFixedHeight(44)
        bar_lay = QHBoxLayout(bar)
        bar_lay.setContentsMargins(12, 6, 12, 6)
        bar_lay.setSpacing(8)

        self._title = QLineEdit()
        self._title.setPlaceholderText("Note title…")
        self._title.setStyleSheet("font-size: 13px; font-weight: 600;")
        bar_lay.addWidget(self._title, 1)

        save_btn = QPushButton("Save")
        save_btn.setObjectName("accentBtn")
        save_btn.setFixedHeight(28)
        save_btn.clicked.connect(self._save)
        bar_lay.addWidget(save_btn)
        lay.addWidget(bar)

        self._editor = QTextEdit()
        self._editor.setPlaceholderText(
            "Start writing your note…\n\n"
            "This note is stored in the NexusOS knowledge graph."
        )
        self._editor.setFont(QFont("SF Pro Display, Helvetica Neue, Arial", 13))
        lay.addWidget(self._editor)

    def _load(self):
        self._title.setText(self._node.get("label", ""))
        meta = self._node.get("meta") or {}
        content = meta.get("note_content") or self._node.get("summary", "")
        self._editor.setPlainText(content)

    def _save(self):
        meta = dict(self._node.get("meta") or {})
        meta["note_content"] = self._editor.toPlainText()
        new_label = self._title.text().strip() or self._node["label"]
        update_node(
            self._node["id"],
            label=new_label,
            summary=self._editor.toPlainText()[:250],
            meta=meta,
        )
        self._node["label"] = new_label
